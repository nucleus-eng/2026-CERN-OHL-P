"""
BCECF pH calibration analysis functions.

Parses Biotek Synergy Neo2 Excel exports (endpoint and kinetic),
calculates 490/440 fluorescence ratios, builds a calibration curve,
and converts kinetic ratio data to pH.
Written by b.next
"""

import datetime
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from scipy.optimize import minimize
from sklearn.metrics import r2_score
from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def load_biotek_endpoint(excel_path: str) -> pd.DataFrame:
    """Parse a Biotek endpoint Excel file with dual-filter (440/490) reads.

    The file has two rows per plate row: first 440nm, then 490nm.
    Column B holds the row letter (on the 440nm line only).
    Column O (index 15) holds the filter code (440535 or 490535).
    Columns C–N (indices 3–14) hold values for plate columns 1–12.

    Returns DataFrame with columns: well, row, column, value_440, value_490
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    records = []
    current_row_letter = None

    # Plate data lives in spreadsheet rows 33–48 (8 plate rows × 2 wavelengths)
    for r in range(33, 49):
        row_letter = ws.cell(r, 2).value  # column B
        filter_code = ws.cell(r, 15).value  # column O

        if row_letter is not None:
            current_row_letter = str(row_letter).strip()

        values = [ws.cell(r, c).value for c in range(3, 15)]  # columns C–N

        if filter_code == 440535:
            for col_idx, val in enumerate(values, start=1):
                records.append({
                    "row": current_row_letter,
                    "column": col_idx,
                    "well": f"{current_row_letter}{col_idx}",
                    "value_440": val if val is not None else 0,
                })
        elif filter_code == 490535:
            for col_idx, val in enumerate(values, start=1):
                well = f"{current_row_letter}{col_idx}"
                # Find the matching 440 record and add 490 value
                for rec in reversed(records):
                    if rec["well"] == well and "value_490" not in rec:
                        rec["value_490"] = val if val is not None else 0
                        break

    df = pd.DataFrame(records)
    return df


def load_biotek_kinetic(excel_path: str, wells: list[str] | None = None) -> pd.DataFrame:
    """Parse a Biotek kinetic Excel file with dual-filter (440/490) reads.

    The file has two sections separated by a marker row:
    - 440nm section: marker '440535' then headers then data
    - 490nm section: marker '490535' then headers then data

    Parameters
    ----------
    excel_path : str
        Path to the Excel file.
    wells : list of str, optional
        If provided, only return data for these wells (e.g. ['C4', 'C5']).

    Returns DataFrame with columns: time_min, well, value_440, value_490
    """
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    max_row = ws.max_row
    max_col = ws.max_column

    # Find section markers and header rows
    marker_rows = []
    for r in range(1, max_row + 1):
        val = ws.cell(r, 1).value
        if val in (440535, 490535):
            marker_rows.append((r, val))

    if len(marker_rows) < 2:
        raise ValueError("Could not find both 440nm and 490nm sections")

    sections = {}
    for marker_row, marker_val in marker_rows:
        # Header row is 2 rows after marker
        header_row = marker_row + 2
        headers = [ws.cell(header_row, c).value for c in range(1, max_col + 1)]

        # Find well columns
        well_cols = {}
        for c_idx, h in enumerate(headers):
            if h and isinstance(h, str) and len(h) <= 3 and h[0].isalpha() and h[1:].isdigit():
                if wells is None or h in wells:
                    well_cols[h] = c_idx

        # Read data rows
        data_rows = []
        for r in range(header_row + 1, max_row + 1):
            time_val = ws.cell(r, 2).value
            if time_val is None:
                break
            # Convert time to minutes
            if isinstance(time_val, datetime.time):
                time_min = time_val.hour * 60 + time_val.minute + time_val.second / 60
            elif isinstance(time_val, (int, float)):
                time_min = float(time_val) * 24 * 60  # days to minutes
            else:
                continue

            for well_name, c_idx in well_cols.items():
                val = ws.cell(r, c_idx + 1).value  # +1 because headers are 0-indexed
                if val is not None:
                    data_rows.append({"time_min": time_min, "well": well_name, "value": val})

        nm_key = "value_440" if marker_val == 440535 else "value_490"
        sections[nm_key] = pd.DataFrame(data_rows)

    # Merge the two sections on time_min and well
    df_440 = sections["value_440"].rename(columns={"value": "value_440"})
    df_490 = sections["value_490"].rename(columns={"value": "value_490"})

    # The 490 timepoints are offset by ~20 seconds; merge on index position
    df_440 = df_440.sort_values(["well", "time_min"]).reset_index(drop=True)
    df_490 = df_490.sort_values(["well", "time_min"]).reset_index(drop=True)

    # Group by well and merge by position within each well
    merged_records = []
    for well_name in df_440["well"].unique():
        w440 = df_440[df_440["well"] == well_name].reset_index(drop=True)
        w490 = df_490[df_490["well"] == well_name].reset_index(drop=True)
        n = min(len(w440), len(w490))
        for i in range(n):
            merged_records.append({
                "time_min": w440.loc[i, "time_min"],
                "well": well_name,
                "value_440": w440.loc[i, "value_440"],
                "value_490": w490.loc[i, "value_490"],
            })

    return pd.DataFrame(merged_records)


def load_platemap(csv_path: str) -> pd.DataFrame:
    """Read a platemap CSV with columns: well, pH (or other metadata)."""
    return pd.read_csv(csv_path)


# ---------------------------------------------------------------------------
# Ratio calculation
# ---------------------------------------------------------------------------

def calculate_ratio(df: pd.DataFrame, numerator: str = "value_490", denominator: str = "value_440") -> pd.DataFrame:
    """Add a 'ratio' column (numerator/denominator) to the DataFrame. Default: 490/440."""
    df = df.copy()
    df["ratio"] = df[numerator] / df[denominator]
    return df


# ---------------------------------------------------------------------------
# Statistics
# ---------------------------------------------------------------------------

def calculate_stats(df: pd.DataFrame, group_col: str = "pH") -> pd.DataFrame:
    """Compute mean and std of ratio grouped by a column (e.g., pH)."""
    return (
        df.groupby(group_col)["ratio"]
        .agg(["mean", "std", "count"])
        .reset_index()
    )


# ---------------------------------------------------------------------------
# Calibration fit
# ---------------------------------------------------------------------------

def fit_calibration(stats_df: pd.DataFrame) -> tuple:
    """Linear regression of ratio vs pH.

    Returns (slope, intercept, r_value, p_value, std_err).
    """
    result = stats.linregress(stats_df["pH"], stats_df["mean"])
    return result


def predict_pH(ratio, slope: float, intercept: float) -> float | np.ndarray:
    """Convert a 490/440 ratio (or array) to pH using the linear calibration.

    pH = (ratio - intercept) / slope
    """
    return (ratio - intercept) / slope


# ---------------------------------------------------------------------------
# Sigmoid curve fitting (kinetic analysis)
# ---------------------------------------------------------------------------

def _sigmoid(t, L, k, t0):
    """Standard logistic: L / (1 + exp(-k*(t - t0)))."""
    return L / (1 + np.exp(-k * (t - t0)))


def _sigmoid_drift(t, L, k, t0, b, tau):
    """Logistic + linear drift: sigmoid(t) + b*(t - tau)."""
    return _sigmoid(t, L, k, t0) + b * (t - tau)


_FIT_FUNCTIONS = {"sigmoid": _sigmoid, "sigmoid_drift": _sigmoid_drift}


def _l2_loss(params, t, y, fit_type):
    fn = _FIT_FUNCTIONS[fit_type]
    return np.sum((y - fn(t, *params)) ** 2)


def _logistic_time_to_fraction(L, k, t0, alpha=0.95):
    """Time when sigmoid reaches *alpha* fraction of asymptote L."""
    if k == 0:
        return np.nan
    return t0 - (1.0 / k) * np.log((1.0 / alpha) - 1.0)


def fit_kinetics(
    kinetic_df: pd.DataFrame,
    cal_slope: float,
    cal_intercept: float,
    fit_type: str = "sigmoid_drift",
    window: int = 3,
) -> pd.DataFrame:
    """Fit a sigmoid (or sigmoid_drift) to each well's pH time-series.

    Parameters
    ----------
    kinetic_df : DataFrame
        Must have columns: time_min, well, ratio.
    cal_slope, cal_intercept : float
        From the linear calibration (ratio = slope*pH + intercept).
    fit_type : str
        'sigmoid' or 'sigmoid_drift'.
    window : int
        Rolling-mean window for initial-guess estimation.

    Returns
    -------
    DataFrame indexed by well with columns:
    vmax, vmax_time_hr, vmax_pH,
    lag_time_hr, lag_pH,
    steady_state_time_hr, steady_state_pH,
    fit_params, r_squared, drift
    """
    df = kinetic_df.copy()
    df["pH"] = predict_pH(df["ratio"], cal_slope, cal_intercept)
    df["time_hr"] = df["time_min"] / 60

    fit_fn = _FIT_FUNCTIONS[fit_type]
    results = []

    for well, wdf in df.groupby("well"):
        wdf = wdf.sort_values("time_hr").reset_index(drop=True)
        t = wdf["time_hr"].values
        y = wdf["pH"].values

        # Smoothed series for initial guesses
        y_smooth = pd.Series(y).rolling(window, min_periods=1).mean().values
        dy = np.diff(y_smooth)
        dt_step = np.median(np.diff(t))

        L0 = float(np.nanmax(y) - np.nanmin(y))
        if L0 == 0:
            L0 = 1.0
        idx_max_slope = int(np.argmax(np.abs(dy)))
        k0 = abs(dy[idx_max_slope]) / dt_step / L0 * 4
        t0_0 = t[min(idx_max_slope, len(t) - 1)]
        offset = float(np.nanmin(y))

        # Shift y so sigmoid starts near 0 (fit L as amplitude, add offset back later)
        y_shifted = y - offset

        p0 = [L0, max(k0, 0.01), t0_0]
        bounds = [(0, L0 * 5), (0, 1000), (0, t.max())]

        if fit_type == "sigmoid_drift":
            p0 += [0.0, t0_0]
            bounds += [(-np.inf, np.inf), (0, t.max())]

        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            try:
                res = minimize(
                    _l2_loss, x0=p0, args=(t, y_shifted, fit_type),
                    method="L-BFGS-B", bounds=bounds,
                )
                params = res.x
            except Exception:
                params = np.array(p0)

        y_fit_shifted = fit_fn(t, *params)
        y_fit = y_fit_shifted + offset
        r2 = r2_score(y, y_fit)

        L, k, t0_fit = params[0], params[1], params[2]
        drift = params[3] if fit_type == "sigmoid_drift" else 0.0

        # Vmax: max slope of the logistic = k * L / 4, occurring at t0
        vmax = k * L / 4
        vmax_time = t0_fit
        vmax_pH = float(fit_fn(t0_fit, *params) + offset)

        # Lag: tangent at inflection intersects y=0 → t_lag = t0 - (L/2) / vmax
        if vmax > 0:
            lag_time = t0_fit - (L / 2) / vmax
            lag_pH = float(fit_fn(max(lag_time, 0), *params) + offset)
        else:
            lag_time = np.nan
            lag_pH = np.nan

        # Steady state: time to 95% of asymptote
        ss_time = _logistic_time_to_fraction(L, k, t0_fit, alpha=0.95)
        ss_pH = float(fit_fn(ss_time, *params) + offset) if np.isfinite(ss_time) else np.nan

        results.append({
            "well": well,
            "vmax": vmax,
            "vmax_time_hr": vmax_time,
            "vmax_pH": vmax_pH,
            "lag_time_hr": lag_time,
            "lag_pH": lag_pH,
            "steady_state_time_hr": ss_time,
            "steady_state_pH": ss_pH,
            "fit_params": params.tolist(),
            "fit_offset": offset,
            "r_squared": r2,
            "drift": drift,
            "fit_type": fit_type,
        })

    return pd.DataFrame(results).set_index("well")


def get_fit_curve(time_hr: np.ndarray, kinetics_row: pd.Series) -> np.ndarray:
    """Evaluate the fitted sigmoid for a single well over an array of times.

    Parameters
    ----------
    time_hr : array of time values (hours)
    kinetics_row : a single row from the fit_kinetics result DataFrame

    Returns
    -------
    array of fitted pH values
    """
    fit_fn = _FIT_FUNCTIONS[kinetics_row["fit_type"]]
    params = kinetics_row["fit_params"]
    offset = kinetics_row["fit_offset"]
    return fit_fn(time_hr, *params) + offset


# ---------------------------------------------------------------------------
# Plotting
# ---------------------------------------------------------------------------

def plot_calibration_curve(
    stats_df: pd.DataFrame,
    fit_result,
    individual_points: pd.DataFrame | None = None,
    ratio_label: str = "490/440",
    ax=None,
):
    """Plot pH calibration curve with error bars and linear regression.

    Parameters
    ----------
    stats_df : DataFrame with columns pH, mean, std
    fit_result : result from scipy.stats.linregress
    individual_points : optional DataFrame with columns pH, ratio (for scatter overlay)
    ax : matplotlib Axes, optional
    """
    if ax is None:
        fig, ax = plt.subplots(figsize=(6, 4))

    # Error bars for mean ± std
    ax.errorbar(
        stats_df["pH"], stats_df["mean"], yerr=stats_df["std"],
        fmt="o", capsize=4, color="steelblue", markersize=8, label="Mean ± SD",
    )

    # Individual data points
    if individual_points is not None:
        ax.scatter(
            individual_points["pH"], individual_points["ratio"],
            alpha=0.5, color="gray", s=30, zorder=1, label="Replicates",
        )

    # Regression line
    x_line = np.linspace(stats_df["pH"].min() - 0.2, stats_df["pH"].max() + 0.2, 100)
    y_line = fit_result.slope * x_line + fit_result.intercept
    ax.plot(x_line, y_line, "--", color="tomato", linewidth=1.5, label="Linear fit")

    # Annotation
    ax.text(
        0.05, 0.95,
        f"y = {fit_result.slope:.4f}x + {fit_result.intercept:.4f}\n$R^2$ = {fit_result.rvalue**2:.4f}",
        transform=ax.transAxes, va="top", fontsize=9,
        bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5),
    )

    ax.set_xlabel("pH")
    ax.set_ylabel(f"Ratio ({ratio_label})")
    ax.set_title("BCECF pH Calibration Curve")
    ax.legend(loc="lower left", fontsize=8)
    sns.despine(ax=ax)
    plt.tight_layout()
    return ax


def plot_kinetic_pH(
    kinetic_df: pd.DataFrame,
    slope: float,
    intercept: float,
    kinetics: pd.DataFrame | None = None,
    show_fit: bool = True,
    annotate: bool = True,
    ax=None,
):
    """Plot pH vs time for kinetic data, converting ratios via calibration.

    Parameters
    ----------
    kinetic_df : DataFrame with columns time_min, well, ratio
    slope, intercept : from linear calibration fit
    kinetics : optional DataFrame from fit_kinetics(); if provided and
        show_fit=True, overlays sigmoid fits and annotates parameters.
    show_fit : bool
        Overlay fitted sigmoid curves (requires kinetics).
    annotate : bool
        Add Vmax / lag / steady-state annotations (requires kinetics).
    ax : matplotlib Axes, optional
    """
    if ax is None:
        fig, ax = plt.subplots(figsize=(8, 4))

    df = kinetic_df.copy()
    df["pH_predicted"] = predict_pH(df["ratio"], slope, intercept)
    df["time_hr"] = df["time_min"] / 60

    colors = sns.color_palette("tab10", n_colors=df["well"].nunique())

    for i, (well, wdf) in enumerate(df.groupby("well")):
        color = colors[i]
        ax.plot(wdf["time_hr"], wdf["pH_predicted"], color=color,
                alpha=0.4, linewidth=1, label=well)

        if kinetics is not None and well in kinetics.index and show_fit:
            row = kinetics.loc[well]
            t_smooth = np.linspace(wdf["time_hr"].min(), wdf["time_hr"].max(), 500)
            y_fit = get_fit_curve(t_smooth, row)
            ax.plot(t_smooth, y_fit, color=color, linewidth=2, linestyle="--")

            if annotate:
                # Steady-state horizontal line
                if np.isfinite(row["steady_state_pH"]):
                    ax.axhline(row["steady_state_pH"], color=color,
                                linewidth=0.5, linestyle=":", alpha=0.5)

    ax.set_xlabel("Time (hours)")
    ax.set_ylabel("Predicted pH")
    ax.set_title("pH Over Time (from BCECF calibration)")
    ax.legend(fontsize=8)
    sns.despine(ax=ax)
    plt.tight_layout()
    return ax


def plot_kinetics_grid(
    kinetic_df: pd.DataFrame,
    slope: float,
    intercept: float,
    kinetics: pd.DataFrame,
):
    """One subplot per well showing data, fit, and annotated kinetic parameters.

    Parameters
    ----------
    kinetic_df : DataFrame with time_min, well, ratio
    slope, intercept : calibration parameters
    kinetics : DataFrame from fit_kinetics()

    Returns
    -------
    matplotlib Figure
    """
    df = kinetic_df.copy()
    df["pH"] = predict_pH(df["ratio"], slope, intercept)
    df["time_hr"] = df["time_min"] / 60

    wells = sorted(df["well"].unique())
    n = len(wells)
    ncols = min(n, 3)
    nrows = -(-n // ncols)  # ceiling division

    fig, axes = plt.subplots(nrows, ncols, figsize=(5 * ncols, 4 * nrows), squeeze=False)
    colors = sns.color_palette("Set2")

    for idx, well in enumerate(wells):
        ax = axes[idx // ncols][idx % ncols]
        wdf = df[df["well"] == well].sort_values("time_hr")
        row = kinetics.loc[well]

        # Data
        ax.scatter(wdf["time_hr"], wdf["pH"], s=6, alpha=0.3, color=colors[2], label="Data")

        # Fit
        t_smooth = np.linspace(wdf["time_hr"].min(), wdf["time_hr"].max(), 500)
        y_fit = get_fit_curve(t_smooth, row)
        ax.plot(t_smooth, y_fit, color=colors[3], linewidth=2, linestyle="--", label="Fit")

        # Vmax tangent line
        vmax = row["vmax"]
        vmax_t = row["vmax_time_hr"]
        vmax_y = row["vmax_pH"]
        tangent_t = wdf["time_hr"].values
        tangent_y = vmax * (tangent_t - vmax_t) + vmax_y
        if row["fit_type"] == "sigmoid_drift":
            params = row["fit_params"]
            tangent_y += params[3] * (tangent_t - params[4])
        mask = (tangent_y > wdf["pH"].min()) & (tangent_y < wdf["pH"].max())
        ax.plot(tangent_t[mask], tangent_y[mask], color=colors[1],
                linewidth=1.5, linestyle="--", label=f"Vmax={vmax:.3f} pH/hr")

        # Steady-state line
        if np.isfinite(row["steady_state_pH"]):
            ax.axhline(row["steady_state_pH"], color=colors[3],
                        linewidth=0.8, linestyle=":", alpha=0.7)
        if np.isfinite(row["steady_state_time_hr"]):
            ax.axvline(row["steady_state_time_hr"], color=colors[3],
                        linewidth=0.8, linestyle=":", alpha=0.7)

        # Annotation box
        txt = (
            f"$V_{{max}}$ = {vmax:.3f} pH/hr\n"
            f"Lag = {row['lag_time_hr']:.1f} hr\n"
            f"SS = {row['steady_state_pH']:.2f} pH\n"
            f"$R^2$ = {row['r_squared']:.3f}"
        )
        ax.text(0.98, 0.05, txt, transform=ax.transAxes, fontsize=8,
                va="bottom", ha="right",
                bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5))

        ax.set_title(well)
        ax.set_xlabel("Time (hours)")
        ax.set_ylabel("Predicted pH")
        ax.legend(fontsize=7, loc="upper left")
        sns.despine(ax=ax)

    # Hide unused axes
    for idx in range(n, nrows * ncols):
        axes[idx // ncols][idx % ncols].set_visible(False)

    fig.tight_layout()
    return fig
